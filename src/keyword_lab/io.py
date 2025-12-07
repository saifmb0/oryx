import json
import logging
import os
from pathlib import Path
from typing import List, Dict, Optional
import pandas as pd

# Try to import openpyxl for Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def write_json(items: List[Dict], output_path: str) -> None:
    if output_path == "-":
        # print formatted JSON array to stdout only
        print(json.dumps(items, ensure_ascii=False, indent=2))
        return
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)
        f.write("\n")
    logging.info(f"Saved JSON to {path}")


def write_csv(items: List[Dict], csv_path: str) -> None:
    if not csv_path:
        return
    path = Path(csv_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(items)
    df.to_csv(path, index=False)
    logging.info(f"Saved CSV to {path}")


def write_excel(items: List[Dict], xlsx_path: str, geo: str = "global") -> None:
    """
    Write results to a professional Excel workbook with multiple sheets.
    
    Creates sheets:
    - 'Keywords': Raw data with all keywords (styled)
    - 'Cluster Summary': Pivot-style summary by cluster
    - 'Intent Analysis': Breakdown by search intent
    - 'Priority Matrix': High-opportunity keywords for action
    - 'Location Analysis': Geographic entity breakdown (UAE only)
    
    Requires openpyxl: pip install keyword-lab[excel]
    
    Args:
        items: List of keyword result dictionaries
        xlsx_path: Path to output .xlsx file
        geo: Geographic target (enables location analysis for 'ae')
    """
    if not xlsx_path:
        return
    
    if not HAS_OPENPYXL:
        logging.warning(
            "Excel export requires openpyxl. Install with: pip install keyword-lab[excel]"
        )
        return
    
    path = Path(xlsx_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    
    df = pd.DataFrame(items)
    
    # Create cluster summary
    if not df.empty and 'cluster' in df.columns:
        summary = df.groupby('cluster').agg({
            'keyword': 'count',
            'search_volume': 'mean',
            'difficulty': 'mean',
            'opportunity_score': 'mean',
        }).round(3)
        summary.columns = ['keyword_count', 'avg_volume', 'avg_difficulty', 'avg_opportunity']
        summary = summary.sort_values('avg_opportunity', ascending=False)
    else:
        summary = pd.DataFrame()
    
    # Create intent analysis
    if not df.empty and 'intent' in df.columns:
        intent_analysis = df.groupby('intent').agg({
            'keyword': 'count',
            'search_volume': 'mean',
            'difficulty': 'mean',
            'opportunity_score': 'mean',
            'ctr_potential': 'mean' if 'ctr_potential' in df.columns else lambda x: 1.0,
        }).round(3)
        intent_analysis.columns = ['keyword_count', 'avg_volume', 'avg_difficulty', 'avg_opportunity', 'avg_ctr']
        intent_analysis = intent_analysis.sort_values('keyword_count', ascending=False)
    else:
        intent_analysis = pd.DataFrame()
    
    # Create priority matrix (high opportunity, lower difficulty)
    if not df.empty:
        priority = df[
            (df['opportunity_score'] >= 0.3) & 
            (df['difficulty'] <= 0.7)
        ].sort_values('opportunity_score', ascending=False).head(50)
    else:
        priority = pd.DataFrame()
    
    # Create location analysis for UAE
    location_analysis = pd.DataFrame()
    if not df.empty and geo.lower() == 'ae':
        try:
            from .entities import extract_entities, UAE_EMIRATES
            
            location_data = []
            for _, row in df.iterrows():
                entities = extract_entities(row['keyword'], 'ae')
                location_data.append({
                    'keyword': row['keyword'],
                    'emirate': entities.get('emirate', 'Not specified'),
                    'district': entities.get('district', ''),
                    'is_local': entities.get('is_local', False),
                    'opportunity_score': row.get('opportunity_score', 0),
                })
            
            location_df = pd.DataFrame(location_data)
            location_analysis = location_df.groupby('emirate').agg({
                'keyword': 'count',
                'is_local': 'sum',
                'opportunity_score': 'mean',
            }).round(3)
            location_analysis.columns = ['keyword_count', 'local_keywords', 'avg_opportunity']
            location_analysis = location_analysis.sort_values('keyword_count', ascending=False)
        except ImportError:
            pass
    
    # Write to Excel with multiple sheets
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        # Main keywords sheet
        df.to_excel(writer, sheet_name='Keywords', index=False)
        
        # Cluster summary
        if not summary.empty:
            summary.to_excel(writer, sheet_name='Cluster Summary')
        
        # Intent analysis
        if not intent_analysis.empty:
            intent_analysis.to_excel(writer, sheet_name='Intent Analysis')
        
        # Priority matrix
        if not priority.empty:
            priority.to_excel(writer, sheet_name='Priority Matrix', index=False)
        
        # Location analysis (UAE only)
        if not location_analysis.empty:
            location_analysis.to_excel(writer, sheet_name='Location Analysis')
        
        # Apply styling
        _style_excel_workbook(writer)
    
    logging.info(f"Saved Excel to {path}")


def _style_excel_workbook(writer: pd.ExcelWriter) -> None:
    """Apply professional styling to Excel workbook."""
    if not HAS_OPENPYXL:
        return
    
    workbook = writer.book
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Style headers
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(50, max(12, max_length + 2))
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze first row
        sheet.freeze_panes = 'A2'


def write_output(items: List[Dict], output_path: str, save_csv: str = None) -> None:
    """
    Write output in the appropriate format based on file extension.
    
    Supports: .json, .csv, .xlsx
    
    Args:
        items: List of keyword result dictionaries
        output_path: Path to output file (or '-' for stdout)
        save_csv: Optional additional CSV path
    """
    if output_path == "-":
        write_json(items, output_path)
        return
    
    path = Path(output_path)
    ext = path.suffix.lower()
    
    if ext == ".xlsx":
        write_excel(items, output_path)
    elif ext == ".csv":
        write_csv(items, output_path)
    else:
        # Default to JSON
        write_json(items, output_path)
    
    # Also write CSV if requested
    if save_csv:
        write_csv(items, save_csv)
