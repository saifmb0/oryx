"""
Input/Output utilities for ORYX (Keyword Lab).

Provides professional Excel reporting with charts, multi-tab analysis,
and stakeholder-ready outputs for the UAE contracting sector.
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd

# Try to import openpyxl for Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None  # <--- FIX 1: Define placeholder to prevent NameError


# =============================================================================
# Color Palette (ORYX Theme)
# =============================================================================

ORYX_COLORS = {
    "primary": "1B5E20",      # Dark green
    "secondary": "2E7D32",    # Medium green
    "accent": "4CAF50",       # Light green
    "success": "66BB6A",      # Success green
    "warning": "FF9800",      # Orange
    "danger": "F44336",       # Red
    "info": "2196F3",         # Blue
    "text": "212121",         # Dark gray
    "light": "FAFAFA",        # Light gray
    "gold": "FFD700",         # Gold for highlights
}


# =============================================================================
# Basic Writers
# =============================================================================

def write_json(items: List[Dict], output_path: str) -> None:
    """Write items to JSON file or stdout."""
    if output_path == "-":
        print(json.dumps(items, ensure_ascii=False, indent=2))
        return
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)
        f.write("\n")
    logging.info(f"Saved JSON to {path}")


def write_csv(items: List[Dict], csv_path: str) -> None:
    """Write items to CSV file."""
    if not csv_path:
        return
    path = Path(csv_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(items)
    df.to_csv(path, index=False)
    logging.info(f"Saved CSV to {path}")


# =============================================================================
# Professional Excel Export (ORYX Edition)
# =============================================================================

def write_excel(
    items: List[Dict], 
    xlsx_path: str, 
    geo: str = "ae",
    report_title: str = "ORYX Keyword Intelligence Report",
    include_charts: bool = True,
    include_executive_summary: bool = True,
) -> None:
    """
    Write results to a professional ORYX-styled Excel workbook.
    """
    if not xlsx_path:
        return
    
    if not HAS_OPENPYXL:
        logging.warning(
            "Excel export requires openpyxl. Install with: pip install keyword-lab[excel]"
        )
        return
    
    path = Path(xlsx_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    
    df = pd.DataFrame(items)
    
    if df.empty:
        logging.warning("No data to export to Excel")
        return
    
    # Create workbook
    workbook = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    
    # 1. Executive Summary
    if include_executive_summary:
        _create_executive_summary(workbook, df, report_title, geo)
    
    # 2. Site Architecture (hierarchical: Parent Topic -> Cluster -> Keywords)
    _create_site_architecture_sheet(workbook, df)
    
    # 3. Keywords Sheet (main data)
    _create_keywords_sheet(workbook, df)
    
    # 4. Cluster Analysis
    _create_cluster_analysis(workbook, df, include_charts)
    
    # 5. Intent Breakdown
    _create_intent_analysis(workbook, df, include_charts)
    
    # 6. Priority Matrix
    _create_priority_matrix(workbook, df)
    
    # 7. GEO Analysis (if available)
    _create_geo_analysis(workbook, df)
    
    # 8. Location Analysis (UAE)
    if geo.lower() == 'ae':
        _create_location_analysis(workbook, df)
    
    # 9. Recommendations
    _create_recommendations(workbook, df, geo)
    
    # Save workbook
    workbook.save(path)
    logging.info(f"Saved professional Excel report to {path}")


def _create_executive_summary(
    workbook: "openpyxl.Workbook",  # <--- FIX 2: Quoted type hint
    df: pd.DataFrame,
    report_title: str,
    geo: str,
) -> None:
    """Create executive summary sheet."""
    sheet = workbook.create_sheet("Executive Summary", 0)
    
    # Title styling
    sheet.merge_cells('A1:E1')
    sheet['A1'] = report_title
    sheet['A1'].font = Font(size=24, bold=True, color=ORYX_COLORS["primary"])
    sheet['A1'].alignment = Alignment(horizontal="center")
    
    # Subtitle with date
    sheet.merge_cells('A2:E2')
    sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Region: {geo.upper()}"
    sheet['A2'].font = Font(size=12, italic=True, color=ORYX_COLORS["text"])
    sheet['A2'].alignment = Alignment(horizontal="center")
    
    # Key Metrics Section
    sheet['A4'] = "KEY METRICS"
    sheet['A4'].font = Font(size=14, bold=True, color=ORYX_COLORS["secondary"])
    
    metrics = [
        ("Total Keywords", len(df)),
        ("Unique Clusters", df['cluster'].nunique() if 'cluster' in df.columns else "N/A"),
        ("Avg. Relative Interest", f"{df['relative_interest'].mean():.2f}" if 'relative_interest' in df.columns else "N/A"),
        ("Avg. Difficulty", f"{df['difficulty'].mean():.2f}" if 'difficulty' in df.columns else "N/A"),
        ("Avg. Opportunity Score", f"{df['opportunity_score'].mean():.2f}" if 'opportunity_score' in df.columns else "N/A"),
        ("High-Priority Keywords", len(df[(df.get('opportunity_score', 0) >= 0.5)]) if 'opportunity_score' in df.columns else "N/A"),
    ]
    
    for i, (label, value) in enumerate(metrics, start=5):
        sheet[f'A{i}'] = label
        sheet[f'A{i}'].font = Font(bold=True)
        sheet[f'B{i}'] = value
        sheet[f'B{i}'].font = Font(color=ORYX_COLORS["primary"])
    
    # Intent Distribution
    if 'intent' in df.columns:
        sheet['A12'] = "INTENT DISTRIBUTION"
        sheet['A12'].font = Font(size=14, bold=True, color=ORYX_COLORS["secondary"])
        
        intent_counts = df['intent'].value_counts()
        for i, (intent, count) in enumerate(intent_counts.items(), start=13):
            pct = (count / len(df)) * 100
            sheet[f'A{i}'] = intent.title()
            sheet[f'B{i}'] = count
            sheet[f'C{i}'] = f"{pct:.1f}%"
    
    # Top Clusters Preview
    if 'cluster' in df.columns:
        sheet['A20'] = "TOP CLUSTERS BY OPPORTUNITY"
        sheet['A20'].font = Font(size=14, bold=True, color=ORYX_COLORS["secondary"])
        
        cluster_opp = df.groupby('cluster')['opportunity_score'].mean().sort_values(ascending=False).head(5)
        for i, (cluster, opp) in enumerate(cluster_opp.items(), start=21):
            sheet[f'A{i}'] = cluster
            sheet[f'B{i}'] = f"{opp:.3f}"
    
    # Column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 15


def _create_site_architecture_sheet(workbook: "openpyxl.Workbook", df: pd.DataFrame) -> None:
    """
    Create a hierarchical site architecture sheet.
    
    Groups keywords by Parent Topic -> Cluster to visualize website structure.
    This helps content managers understand the recommended site hierarchy.
    """
    if 'parent_topic' not in df.columns or 'cluster' not in df.columns:
        return
    
    sheet = workbook.create_sheet("Site Architecture")
    
    # Header
    sheet['A1'] = "SITE ARCHITECTURE"
    sheet['A1'].font = Font(size=16, bold=True, color=ORYX_COLORS["primary"])
    sheet.merge_cells('A1:D1')
    
    sheet['A2'] = "Hierarchical view: Parent Topic → Cluster → Keywords"
    sheet['A2'].font = Font(size=10, italic=True, color="666666")
    
    # Column headers
    headers = ["Level", "Topic/Cluster/Keyword", "Intent", "Opportunity"]
    for c_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=c_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=ORYX_COLORS["primary"], 
                                end_color=ORYX_COLORS["primary"], fill_type="solid")
    
    row_idx = 5
    
    # Group by parent_topic, then cluster
    for parent_topic in df['parent_topic'].unique():
        # L1: Parent Topic (Pillar)
        cell = sheet.cell(row=row_idx, column=1, value="L1 - Pillar")
        cell.font = Font(bold=True, color=ORYX_COLORS["primary"])
        
        cell = sheet.cell(row=row_idx, column=2, value=parent_topic.title())
        cell.font = Font(bold=True, size=12)
        
        # Calculate average opportunity for parent topic
        pt_df = df[df['parent_topic'] == parent_topic]
        avg_opp = pt_df['opportunity_score'].mean()
        sheet.cell(row=row_idx, column=4, value=f"{avg_opp:.2f}")
        
        row_idx += 1
        
        # L2: Clusters under this parent topic
        for cluster in pt_df['cluster'].unique():
            cell = sheet.cell(row=row_idx, column=1, value="  L2 - Cluster")
            cell.font = Font(color=ORYX_COLORS["secondary"])
            
            cell = sheet.cell(row=row_idx, column=2, value=f"  └─ {cluster}")
            cell.font = Font(bold=True, color=ORYX_COLORS["secondary"])
            
            cluster_df = pt_df[pt_df['cluster'] == cluster]
            cluster_opp = cluster_df['opportunity_score'].mean()
            sheet.cell(row=row_idx, column=4, value=f"{cluster_opp:.2f}")
            
            row_idx += 1
            
            # L3: Keywords in this cluster
            for _, kw_row in cluster_df.head(5).iterrows():  # Top 5 per cluster
                sheet.cell(row=row_idx, column=1, value="    L3 - Keyword")
                sheet.cell(row=row_idx, column=2, value=f"      └─ {kw_row['keyword']}")
                sheet.cell(row=row_idx, column=3, value=kw_row.get('intent', ''))
                sheet.cell(row=row_idx, column=4, value=f"{kw_row['opportunity_score']:.2f}")
                row_idx += 1
        
        row_idx += 1  # Add spacing between parent topics
    
    # Column widths
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 12
    
    # Freeze header
    sheet.freeze_panes = 'A5'


def _create_keywords_sheet(workbook: "openpyxl.Workbook", df: pd.DataFrame) -> None:
    """Create main keywords data sheet with conditional formatting."""
    sheet = workbook.create_sheet("Keywords")
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                # Header styling
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=ORYX_COLORS["primary"], 
                                        end_color=ORYX_COLORS["primary"], fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
    
    # Apply conditional formatting to opportunity_score column
    if 'opportunity_score' in df.columns:
        opp_col = list(df.columns).index('opportunity_score') + 1
        opp_col_letter = get_column_letter(opp_col)
        
        # Color scale: red -> yellow -> green
        color_scale = ColorScaleRule(
            start_type='num', start_value=0, start_color='F44336',
            mid_type='num', mid_value=0.5, mid_color='FFEB3B',
            end_type='num', end_value=1, end_color='4CAF50'
        )
        sheet.conditional_formatting.add(
            f'{opp_col_letter}2:{opp_col_letter}{len(df)+1}', 
            color_scale
        )
    
    # Apply conditional formatting to difficulty column
    if 'difficulty' in df.columns:
        diff_col = list(df.columns).index('difficulty') + 1
        diff_col_letter = get_column_letter(diff_col)
        
        # Inverse color scale: green -> yellow -> red
        color_scale = ColorScaleRule(
            start_type='num', start_value=0, start_color='4CAF50',
            mid_type='num', mid_value=0.5, mid_color='FFEB3B',
            end_type='num', end_value=1, end_color='F44336'
        )
        sheet.conditional_formatting.add(
            f'{diff_col_letter}2:{diff_col_letter}{len(df)+1}',
            color_scale
        )
    
    # Auto-fit columns
    _auto_fit_columns(sheet)
    
    # Freeze header row
    sheet.freeze_panes = 'A2'


def _create_cluster_analysis(
    workbook: "openpyxl.Workbook", 
    df: pd.DataFrame,
    include_charts: bool,
) -> None:
    """Create cluster analysis sheet with optional charts."""
    if 'cluster' not in df.columns:
        return
    
    sheet = workbook.create_sheet("Cluster Analysis")
    
    # Create summary
    summary = df.groupby('cluster').agg({
        'keyword': 'count',
        'relative_interest': ['sum', 'mean'],
        'difficulty': 'mean',
        'opportunity_score': 'mean',
    }).round(3)
    summary.columns = ['count', 'total_interest', 'avg_interest', 'avg_difficulty', 'avg_opportunity']
    summary = summary.sort_values('avg_opportunity', ascending=False).reset_index()
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(summary, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=ORYX_COLORS["secondary"],
                                        end_color=ORYX_COLORS["secondary"], fill_type="solid")
    
    # Add chart
    if include_charts and len(summary) > 1:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Cluster Opportunity Scores"
        chart.x_axis.title = "Cluster"
        chart.y_axis.title = "Avg Opportunity"
        
        data = Reference(sheet, min_col=6, min_row=1, max_row=len(summary)+1)
        cats = Reference(sheet, min_col=1, min_row=2, max_row=len(summary)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        
        # Color bars green
        for series in chart.series:
            series.graphicalProperties.solidFill = ORYX_COLORS["accent"]
        
        sheet.add_chart(chart, "H2")
    
    _auto_fit_columns(sheet)


def _create_intent_analysis(
    workbook: "openpyxl.Workbook",
    df: pd.DataFrame,
    include_charts: bool,
) -> None:
    """Create intent analysis sheet."""
    if 'intent' not in df.columns:
        return
    
    sheet = workbook.create_sheet("Intent Breakdown")
    
    # Create summary
    intent_summary = df.groupby('intent').agg({
        'keyword': 'count',
        'relative_interest': 'mean',
        'difficulty': 'mean',
        'opportunity_score': 'mean',
    }).round(3)
    intent_summary.columns = ['count', 'avg_interest', 'avg_difficulty', 'avg_opportunity']
    intent_summary = intent_summary.sort_values('count', ascending=False).reset_index()
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(intent_summary, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=ORYX_COLORS["info"],
                                        end_color=ORYX_COLORS["info"], fill_type="solid")
    
    # Add pie chart
    if include_charts and len(intent_summary) > 1:
        chart = PieChart()
        chart.title = "Intent Distribution"
        
        data = Reference(sheet, min_col=2, min_row=2, max_row=len(intent_summary)+1)
        labels = Reference(sheet, min_col=1, min_row=2, max_row=len(intent_summary)+1)
        chart.add_data(data)
        chart.set_categories(labels)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        
        sheet.add_chart(chart, "G2")
    
    _auto_fit_columns(sheet)


def _create_priority_matrix(workbook: "openpyxl.Workbook", df: pd.DataFrame) -> None:
    """Create priority matrix of high-opportunity, low-difficulty keywords."""
    sheet = workbook.create_sheet("Priority Matrix")
    
    # Filter high-priority keywords
    if 'opportunity_score' in df.columns and 'difficulty' in df.columns:
        priority = df[
            (df['opportunity_score'] >= 0.3) &
            (df['difficulty'] <= 0.7)
        ].sort_values('opportunity_score', ascending=False).head(50)
    else:
        priority = df.head(50)
    
    if priority.empty:
        sheet['A1'] = "No high-priority keywords found."
        return
    
    # Header
    sheet.merge_cells('A1:F1')
    sheet['A1'] = "🎯 HIGH-PRIORITY KEYWORDS (Opp ≥ 0.3, Diff ≤ 0.7)"
    sheet['A1'].font = Font(size=14, bold=True, color=ORYX_COLORS["gold"])
    sheet['A1'].fill = PatternFill(start_color=ORYX_COLORS["primary"],
                                    end_color=ORYX_COLORS["primary"], fill_type="solid")
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(priority, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 3:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=ORYX_COLORS["accent"],
                                        end_color=ORYX_COLORS["accent"], fill_type="solid")
    
    _auto_fit_columns(sheet)


def _create_geo_analysis(workbook: "openpyxl.Workbook", df: pd.DataFrame) -> None:
    """Create GEO (Generative Engine Optimization) analysis sheet."""
    # Check for GEO-related columns
    geo_cols = ['geo_suitability', 'info_gain_score', 'geo_query_type']
    has_geo = any(col in df.columns for col in geo_cols)
    
    if not has_geo:
        return
    
    sheet = workbook.create_sheet("GEO Analysis")
    
    # Header
    sheet.merge_cells('A1:F1')
    sheet['A1'] = "🤖 AI SEARCH OPTIMIZATION (GEO Analysis)"
    sheet['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    sheet['A1'].fill = PatternFill(start_color=ORYX_COLORS["info"],
                                    end_color=ORYX_COLORS["info"], fill_type="solid")
    
    # Filter to keywords with GEO data
    geo_df = df[[col for col in df.columns if col in ['keyword', 'cluster'] + geo_cols]]
    
    for r_idx, row in enumerate(dataframe_to_rows(geo_df, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 3:
                cell.font = Font(bold=True)
    
    _auto_fit_columns(sheet)


def _create_location_analysis(workbook: "openpyxl.Workbook", df: pd.DataFrame) -> None:
    """Create UAE location analysis sheet."""
    sheet = workbook.create_sheet("Location Analysis")
    
    # Header
    sheet.merge_cells('A1:E1')
    sheet['A1'] = "📍 UAE LOCATION ANALYSIS"
    sheet['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    sheet['A1'].fill = PatternFill(start_color=ORYX_COLORS["warning"],
                                    end_color=ORYX_COLORS["warning"], fill_type="solid")
    
    try:
        from .entities import extract_entities
        
        location_data = []
        for _, row in df.iterrows():
            entities = extract_entities(row['keyword'], 'ae')
            location_data.append({
                'keyword': row['keyword'],
                'emirate': entities.get('emirate', 'Not specified'),
                'district': entities.get('district', ''),
                'is_local': entities.get('is_local', False),
                'is_contracting': entities.get('is_contracting', False),
                'opportunity_score': row.get('opportunity_score', 0),
            })
        
        loc_df = pd.DataFrame(location_data)
        
        # Summary by emirate
        summary = loc_df.groupby('emirate').agg({
            'keyword': 'count',
            'is_local': 'sum',
            'is_contracting': 'sum',
            'opportunity_score': 'mean',
        }).round(3)
        summary.columns = ['total_keywords', 'local_keywords', 'contracting_keywords', 'avg_opportunity']
        summary = summary.sort_values('total_keywords', ascending=False).reset_index()
        
        # Write summary
        sheet['A3'] = "Summary by Emirate"
        sheet['A3'].font = Font(bold=True, size=12)
        
        for r_idx, row in enumerate(dataframe_to_rows(summary, index=False, header=True), 5):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 5:
                    cell.font = Font(bold=True)
        
    except ImportError:
        sheet['A3'] = "Entity extraction module not available"
    
    _auto_fit_columns(sheet)


def _create_recommendations(
    workbook: "openpyxl.Workbook", 
    df: pd.DataFrame,
    geo: str,
) -> None:
    """Create actionable recommendations sheet."""
    sheet = workbook.create_sheet("Recommendations")
    
    # Header
    sheet.merge_cells('A1:D1')
    sheet['A1'] = "📋 ACTIONABLE RECOMMENDATIONS"
    sheet['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    sheet['A1'].fill = PatternFill(start_color=ORYX_COLORS["success"],
                                    end_color=ORYX_COLORS["success"], fill_type="solid")
    
    recommendations = []
    
    # Analyze data and generate recommendations
    if 'opportunity_score' in df.columns:
        high_opp = df[df['opportunity_score'] >= 0.6]
        if len(high_opp) > 0:
            recommendations.append({
                "priority": "HIGH",
                "category": "Quick Wins",
                "recommendation": f"Prioritize {len(high_opp)} high-opportunity keywords (score ≥ 0.6)",
                "action": "Create or optimize content for these keywords first",
            })
    
    if 'cluster' in df.columns:
        top_cluster = df.groupby('cluster')['opportunity_score'].mean().idxmax()
        recommendations.append({
            "priority": "HIGH",
            "category": "Content Strategy",
            "recommendation": f"Focus content efforts on '{top_cluster}' cluster",
            "action": "Develop comprehensive pillar content for this topic",
        })
    
    if 'intent' in df.columns:
        intent_counts = df['intent'].value_counts()
        top_intent = intent_counts.idxmax()
        recommendations.append({
            "priority": "MEDIUM",
            "category": "Content Format",
            "recommendation": f"Most keywords are {top_intent} intent ({intent_counts[top_intent]} keywords)",
            "action": f"Optimize content format for {top_intent} search intent",
        })
    
    if geo.lower() == 'ae':
        recommendations.append({
            "priority": "MEDIUM",
            "category": "Localization",
            "recommendation": "Target Abu Dhabi-specific terms",
            "action": "Include TAMM, Estidama, and local district names in content",
        })
    
    # Write recommendations
    headers = ["Priority", "Category", "Recommendation", "Action"]
    for c_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=c_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=ORYX_COLORS["light"],
                                end_color=ORYX_COLORS["light"], fill_type="solid")
    
    for r_idx, rec in enumerate(recommendations, 4):
        sheet.cell(row=r_idx, column=1, value=rec["priority"])
        sheet.cell(row=r_idx, column=2, value=rec["category"])
        sheet.cell(row=r_idx, column=3, value=rec["recommendation"])
        sheet.cell(row=r_idx, column=4, value=rec["action"])
        
        # Color priority
        priority_cell = sheet.cell(row=r_idx, column=1)
        if rec["priority"] == "HIGH":
            priority_cell.fill = PatternFill(start_color=ORYX_COLORS["danger"],
                                              end_color=ORYX_COLORS["danger"], fill_type="solid")
            priority_cell.font = Font(bold=True, color="FFFFFF")
        elif rec["priority"] == "MEDIUM":
            priority_cell.fill = PatternFill(start_color=ORYX_COLORS["warning"],
                                              end_color=ORYX_COLORS["warning"], fill_type="solid")
    
    _auto_fit_columns(sheet)


def _auto_fit_columns(sheet) -> None:
    """Auto-fit column widths based on content."""
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(60, max(12, max_length + 2))
        sheet.column_dimensions[column_letter].width = adjusted_width


# =============================================================================
# SEO Metadata Auto-Generation
# =============================================================================

def generate_seo_metadata(
    items: List[Dict],
    brand_name: str = "HAGCC",
    location: str = "Abu Dhabi",
    output_path: Optional[str] = None,
) -> List[Dict]:
    """
    Generate SEO metadata (title tags, meta descriptions) for keyword clusters.
    
    Automatically creates optimized meta tags following SEO best practices:
    - Title: 50-60 characters with primary keyword + brand
    - Description: 150-160 characters with keyword variations
    
    Args:
        items: List of keyword dicts from pipeline
        brand_name: Company/brand name for title suffix
        location: Primary location for local SEO
        output_path: Optional path to save metadata.json
        
    Returns:
        List of metadata dicts for each unique cluster
    """
    # Group keywords by cluster
    clusters: Dict[str, List[Dict]] = {}
    for item in items:
        cluster = item.get("cluster", "general")
        if cluster not in clusters:
            clusters[cluster] = []
        clusters[cluster].append(item)
    
    metadata_list = []
    
    for cluster_name, kws in clusters.items():
        # Sort by opportunity score to get best keywords
        sorted_kws = sorted(kws, key=lambda x: x.get("opportunity_score", 0), reverse=True)
        
        if not sorted_kws:
            continue
        
        # Primary keyword (highest opportunity)
        primary = sorted_kws[0]
        primary_kw = primary.get("keyword", cluster_name).title()
        
        # Secondary keywords for description
        secondary_kws = [kw.get("keyword", "") for kw in sorted_kws[1:4]]
        
        # Get parent topic for context
        parent_topic = primary.get("parent_topic", cluster_name).title()
        
        # Generate title (50-60 chars target)
        # Pattern: "Best {Primary Keyword} | {Brand} {Location}"
        title_base = f"Best {primary_kw}"
        title_suffix = f" | {brand_name} {location}"
        
        if len(title_base + title_suffix) <= 60:
            title = title_base + title_suffix
        else:
            # Truncate if too long
            max_base = 60 - len(title_suffix) - 3
            title = title_base[:max_base] + "..." + title_suffix
        
        # Generate description (150-160 chars target)
        # Pattern: "Expert {Parent Topic} in {Location}. Services: {kw1}, {kw2}, {kw3}. Contact us today!"
        desc_intro = f"Expert {parent_topic} services in {location}."
        
        if secondary_kws:
            services = ", ".join([kw.title() for kw in secondary_kws if kw])
            desc_services = f" We offer: {services}."
        else:
            desc_services = ""
        
        desc_cta = " Contact us for a free quote!"
        
        full_desc = desc_intro + desc_services + desc_cta
        
        # Truncate if over 160 chars
        if len(full_desc) > 160:
            full_desc = full_desc[:157] + "..."
        
        # Intent-based schema type suggestion
        intent = primary.get("intent", "informational")
        schema_type = _suggest_schema_type(intent, cluster_name)
        
        metadata = {
            "cluster": cluster_name,
            "title": title,
            "description": full_desc,
            "primary_keyword": primary.get("keyword", ""),
            "secondary_keywords": secondary_kws,
            "parent_topic": primary.get("parent_topic", ""),
            "intent": intent,
            "suggested_schema": schema_type,
            "h1_suggestion": f"{parent_topic} - {primary_kw}",
            "keywords_count": len(sorted_kws),
            "avg_opportunity": sum(k.get("opportunity_score", 0) for k in sorted_kws) / len(sorted_kws),
        }
        
        metadata_list.append(metadata)
    
    # Sort by average opportunity
    metadata_list = sorted(metadata_list, key=lambda x: x.get("avg_opportunity", 0), reverse=True)
    
    # Save to file if path provided
    if output_path:
        write_json(metadata_list, output_path)
        logging.info(f"Generated SEO metadata for {len(metadata_list)} clusters")
    
    return metadata_list


def _suggest_schema_type(intent: str, cluster_name: str) -> str:
    """Suggest appropriate schema.org type based on intent and cluster."""
    intent_lower = intent.lower()
    cluster_lower = cluster_name.lower()
    
    # Service-based clusters
    if any(term in cluster_lower for term in ["service", "contractor", "company", "repair", "install"]):
        return "LocalBusiness"
    
    # Product-based
    if any(term in cluster_lower for term in ["product", "equipment", "material", "supply"]):
        return "Product"
    
    # FAQ/How-to content
    if intent_lower == "informational":
        return "FAQPage"
    
    # Transactional
    if intent_lower == "transactional":
        return "Service"
    
    # Local
    if intent_lower == "local":
        return "LocalBusiness"
    
    return "WebPage"


# =============================================================================
# Output Router
# =============================================================================

def generate_run_id() -> str:
    """Generate a timestamped run ID in YYYYMMDDHHMM format."""
    return datetime.now().strftime("%Y%m%d%H%M")


def get_run_dir(base_dir: str = "./data", run_id: Optional[str] = None) -> Path:
    """
    Get or create a timestamped run directory.
    
    Args:
        base_dir: Base directory for runs (default: ./data)
        run_id: Optional run ID. If None, generates a new one.
        
    Returns:
        Path to the run directory (e.g., ./data/run_id=202506151430/)
    """
    if run_id is None:
        run_id = generate_run_id()
    run_dir = Path(base_dir) / f"run_id={run_id}"
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir


def write_output(
    items: List[Dict], 
    output_path: str, 
    save_csv: str = None,
    geo: str = "ae",
    report_title: str = "ORYX Keyword Intelligence Report",
    use_run_dir: bool = False,
    run_id: Optional[str] = None,
) -> Optional[Path]:
    """
    Write output in the appropriate format based on file extension.
    
    Args:
        items: List of keyword dictionaries to write.
        output_path: Path to output file (or "-" for stdout).
        save_csv: Optional path for additional CSV export.
        geo: Geographic region for Excel report.
        report_title: Title for Excel report.
        use_run_dir: If True, wrap outputs in ./data/run_id=YYYYMMDDHHMM/
        run_id: Optional run ID. If None and use_run_dir=True, generates one.
        
    Returns:
        Path to the run directory if use_run_dir=True, otherwise None.
    """
    if output_path == "-":
        write_json(items, output_path)
        return None
    
    path = Path(output_path)
    ext = path.suffix.lower()
    
    # Optionally wrap in timestamped run directory
    run_dir = None
    if use_run_dir:
        run_dir = get_run_dir(run_id=run_id)
        path = run_dir / path.name
        logging.info(f"Writing outputs to run directory: {run_dir}")
    
    if ext == ".xlsx":
        write_excel(items, str(path), geo=geo, report_title=report_title)
    elif ext == ".csv":
        write_csv(items, str(path))
    else:
        write_json(items, str(path))
    
    if save_csv:
        csv_path = run_dir / Path(save_csv).name if run_dir else save_csv
        write_csv(items, str(csv_path))
    
    return run_dir